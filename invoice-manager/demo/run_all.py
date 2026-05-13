"""
demo/run_all.py — Full pipeline demo for every fiscal profile.

Generates synthetic PDF invoices, runs extract + export for each profile,
then verifies outputs. Run from project root:

    python3 demo/run_all.py
"""

import io
import sqlite3
import subprocess
import sys
from pathlib import Path

from fpdf import FPDF
from fpdf.enums import XPos, YPos

ROOT = Path(__file__).parent.parent
TESTING = Path(__file__).parent

# ── Synthetic invoice texts ───────────────────────────────────────────────────

def _pdf(lines: list[str]) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    for line in lines:
        pdf.cell(0, 5, line, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()


INVOICES = {
    "auto-entrepreneur": [
        ("facture-hebergement-ae.pdf", [
            "OVH SAS -SIREN 424761419 -N TVA FR22424761419",
            "Facture : AE-2025-001",
            "Date d emission : 01 Mars 2025",
            "Hebergement serveur dedie 12 mois",
            "Total HT : 107,88 EUR",
            "TVA 20% : 21,58 EUR",
            "Total TTC : 129,46 EUR",
            "Paiement par prelevement",
        ]),
        ("facture-logiciel-ae.pdf", [
            "Adobe Systems -SIREN 987654321",
            "Facture : AE-2025-002",
            "Date d emission : 15 Juin 2025",
            "Abonnement Creative Cloud annuel",
            "Total HT : 499,00 EUR",
            "TVA 20% : 99,80 EUR",
            "Total TTC : 598,80 EUR",
            "Paiement par CB",
        ]),
        ("avoir-fournisseur-ae.pdf", [
            "OVH SAS -SIREN 424761419",
            "AVOIR -Note de credit AV-2025-001",
            "Remboursement avoir fournisseur",
            "Date : 20 Juillet 2025",
            "Montant HT : -50,00 EUR",
            "TVA 20% : -10,00 EUR",
            "Montant TTC : -60,00 EUR",
        ]),
        ("recu-fournitures.pdf", [
            "FNAC PARIS",
            "Date : 10 Septembre 2025",
            "Souris sans fil",
            "TOTAL TTC : 29,99 EUR",
            "TVA 20% : 5,00 EUR",
            "HT : 24,99 EUR",
            "Paiement carte bancaire",
        ]),
    ],
    "sasu": [
        ("facture-emise-sasu.pdf", [
            "Tech Solutions SASU -SIREN 222222222 -N TVA FR22222222222",
            "Facture client : SASU-2025-001",
            "Date : 05 Fevrier 2025",
            "Prestation developpement web -10 jours",
            "Client : Dupont SAS",
            "Total HT : 5000,00 EUR",
            "TVA 20% : 1000,00 EUR",
            "Total TTC : 6000,00 EUR",
            "Virement bancaire",
        ]),
        ("facture-loyer-bureau-sasu.pdf", [
            "Fonciere Immo -SIREN 333333333",
            "Facture : LOY-2025-001",
            "Date : 01 Janvier 2025",
            "Loyer bureau Q1 2025",
            "Total HT : 1500,00 EUR",
            "TVA 20% : 300,00 EUR",
            "Total TTC : 1800,00 EUR",
            "Prelevement mensuel",
        ]),
        ("note-de-frais-sasu.pdf", [
            "Note de frais -Remboursement de frais professionnels",
            "Salarie : Martin Dupont",
            "Date : 28 Mars 2025",
            "Deplacement Paris Lyon TGV",
            "Montant TTC : 87,00 EUR",
            "TVA 10% : 7,91 EUR",
            "HT : 79,09 EUR",
            "Carte bancaire personnelle",
        ]),
        ("avoir-client-sasu.pdf", [
            "Tech Solutions SASU -SIREN 222222222",
            "AVOIR CLIENT -AV-SASU-2025-001",
            "Avoir credit note remboursement",
            "Date : 15 Avril 2025",
            "Prestation annulee",
            "Montant HT : -500,00 EUR",
            "TVA 20% : -100,00 EUR",
            "TTC : -600,00 EUR",
        ]),
    ],
    "sarl": [
        ("facture-emise-sarl.pdf", [
            "BTP Solutions SARL -SIREN 444444444 -N TVA FR44444444444",
            "Facture : SARL-2025-001",
            "Date : 10 Mars 2025",
            "Travaux renovation batiment",
            "Client : Mairie de Lyon",
            "Total HT : 12000,00 EUR",
            "TVA 20% : 2400,00 EUR",
            "Total TTC : 14400,00 EUR",
            "Virement",
        ]),
        ("facture-materiel-sarl.pdf", [
            "Leroy Merlin -SIREN 555555555",
            "Facture : LM-2025-4521",
            "Date : 22 Fevrier 2025",
            "Outillage professionnel",
            "Total HT : 850,00 EUR",
            "TVA 20% : 170,00 EUR",
            "Total TTC : 1020,00 EUR",
            "Cheque",
        ]),
        ("avoir-fournisseur-sarl.pdf", [
            "Leroy Merlin -SIREN 555555555",
            "AVOIR -AV-LM-2025-001",
            "Credit note -retour marchandise",
            "Date : 05 Mars 2025",
            "Montant HT : -200,00 EUR",
            "TVA 20% : -40,00 EUR",
            "TTC : -240,00 EUR",
        ]),
        ("note-de-frais-sarl.pdf", [
            "Note de frais remboursement de frais",
            "Employe : Sophie Martin",
            "Date : 18 Avril 2025",
            "Repas client restaurant affaires",
            "Total TTC : 95,00 EUR",
            "TVA 10% : 8,64 EUR",
            "HT : 86,36 EUR",
        ]),
    ],
    "salarie": [
        ("note-de-frais-salarie.pdf", [
            "Note de frais -Remboursement de frais professionnels",
            "Salarie : Pierre Durand",
            "Employeur : Grande Entreprise SA",
            "Date : 12 Janvier 2025",
            "Transport : Paris Bordeaux",
            "Total TTC : 120,00 EUR",
            "TVA 10% : 10,91 EUR",
            "HT : 109,09 EUR",
        ]),
        ("facture-formation-salarie.pdf", [
            "OpenClassrooms -SIREN 666666666",
            "Facture : OC-2025-001",
            "Date : 03 Fevrier 2025",
            "Formation Python avance -CPF",
            "Total HT : 1500,00 EUR",
            "TVA 20% : 300,00 EUR",
            "Total TTC : 1800,00 EUR",
            "Virement OPCO",
        ]),
        ("recu-transport-salarie.pdf", [
            "RATP",
            "Date : 01 Mars 2025",
            "Pass Navigo mensuel",
            "TOTAL : 86,40 EUR",
            "Carte bancaire",
        ]),
    ],
}

CONFIGS = {
    "auto-entrepreneur": {
        "siren": "111111111",
        "cadence": "trimestrielle",
        "profile": "auto-entrepreneur",
    },
    "sasu": {
        "siren": "222222222",
        "cadence": "mensuelle",
        "profile": "SASU",
    },
    "sarl": {
        "siren": "444444444",
        "cadence": "mensuelle",
        "profile": "SARL",
    },
    "salarie": {
        "siren": "",
        "cadence": "annuelle",
        "profile": "salarié",
    },
}

# ── Setup ─────────────────────────────────────────────────────────────────────

def setup_profile(name: str) -> Path:
    profile_dir = TESTING / name
    for d in ("input", "processed", "errors", "data", "output", "review"):
        (profile_dir / d).mkdir(parents=True, exist_ok=True)

    # Generate PDFs
    for filename, lines in INVOICES[name]:
        (profile_dir / "input" / filename).write_bytes(_pdf(lines))

    return profile_dir


# ── Run ───────────────────────────────────────────────────────────────────────

def run(cmd: list[str], cwd: Path) -> tuple[int, str]:
    result = subprocess.run(
        [sys.executable] + cmd,
        cwd=str(cwd),
        capture_output=True,
        text=True,
    )
    return result.returncode, result.stdout + result.stderr


# ── Verify ────────────────────────────────────────────────────────────────────

def verify(name: str, profile_dir: Path) -> list[str]:
    errors = []
    cfg = CONFIGS[name]
    db_path = profile_dir / "data" / "invoices.db"

    if not db_path.exists():
        errors.append("DB manquante")
        return errors

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM invoices").fetchall()

    # Row count
    expected_n = len(INVOICES[name])
    if len(rows) != expected_n:
        errors.append(f"DB : {len(rows)} lignes, attendu {expected_n}")

    # Fiscal profile
    profiles = {r["statut_fiscal_profil"] for r in rows}
    if profiles != {cfg["profile"]}:
        errors.append(f"Profil fiscal inattendu : {profiles}")

    # Document types per profile
    types = {r["type_document"] for r in rows}
    if name in ("sasu", "sarl") and "facture_émise" not in types:
        errors.append(f"facture_émise non détectée (SIREN={cfg['siren']})")
    if name in ("sasu", "sarl", "salarie") and "note_de_frais" not in types:
        errors.append("note_de_frais non détectée")
    if name == "auto-entrepreneur" and "avoir_reçu" not in types:
        errors.append("avoir_reçu non détecté")

    conn.close()

    # XLSX
    import openpyxl
    xlsx_files = list((profile_dir / "output").glob("*.xlsx"))
    if not xlsx_files:
        errors.append("Aucun XLSX généré")
    else:
        for xlsx in xlsx_files:
            wb = openpyxl.load_workbook(xlsx)
            expected_sheets = {"Journal", "Récapitulatif", "Déclaration", "Statistiques"}
            if set(wb.sheetnames) != expected_sheets:
                errors.append(f"{xlsx.name} : onglets {wb.sheetnames}, attendu {expected_sheets}")

    return errors


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Invoice Manager -Test pipeline complet par profil fiscal")
    print("=" * 60)

    all_ok = True

    for name in CONFIGS:
        print(f"\n{'─' * 50}")
        print(f"Profil : {name.upper()}")
        print(f"{'─' * 50}")

        profile_dir = setup_profile(name)
        print(f"  Dossier   : {profile_dir.relative_to(ROOT)}")
        print(f"  Factures  : {len(INVOICES[name])} fichiers générés")

        # Extract
        code, out = run([str(ROOT / "extract.py")], profile_dir)
        for line in out.strip().splitlines():
            print(f"  {line}")
        if code != 0:
            print(f"  [ERREUR] extract.py a échoué (code {code})")
            all_ok = False
            continue

        # Export -all years (--all avoids defaulting to current calendar year)
        code, out = run([str(ROOT / "export.py"), "--all"], profile_dir)
        for line in out.strip().splitlines():
            print(f"  {line}")
        if code != 0:
            print(f"  [ERREUR] export.py a échoué (code {code})")
            all_ok = False
            continue

        # Verify
        errs = verify(name, profile_dir)
        if errs:
            all_ok = False
            for e in errs:
                print(f"  [FAIL] {e}")
        else:
            # Summary from DB
            import sqlite3 as sq
            conn = sq.connect(profile_dir / "data" / "invoices.db")
            conn.row_factory = sq.Row
            rows = conn.execute("SELECT type_document, statut_révision FROM invoices").fetchall()
            conn.close()
            type_counts = {}
            for r in rows:
                type_counts[r["type_document"]] = type_counts.get(r["type_document"], 0) + 1
            rev_counts = {}
            for r in rows:
                rev_counts[r["statut_révision"]] = rev_counts.get(r["statut_révision"], 0) + 1
            print(f"  [OK] {len(rows)} entrées en base")
            for t, n in sorted(type_counts.items()):
                print(f"       · {t} ×{n}")
            for s, n in sorted(rev_counts.items()):
                print(f"       · statut {s} ×{n}")
            outputs = list((profile_dir / "output").glob("*"))
            print(f"  [OK] {len(outputs)} fichiers dans output/")
            for f in sorted(outputs):
                print(f"       · {f.name}")

    print(f"\n{'=' * 60}")
    print("RÉSULTAT :", "TOUT OK ✓" if all_ok else "ÉCHECS DÉTECTÉS ✗")
    print("=" * 60)
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
