"""
export.py — Export CSV + XLSX du carnet de compte
Usage:
  python export.py --year 2025
  python export.py --year 2025 --statut auto-entrepreneur
  python export.py --all
"""

import argparse
import calendar
import csv
import shutil
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

from config import load_config, CADENCE_DEFAULTS
from constants import FISCAL_RULES, MONTHS_FR_SHORT, STATUT_A_REVISER
from db import get_user_profile, open_db

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Requête ───────────────────────────────────────────────────────────────────

def fetch_rows(conn: sqlite3.Connection, year: int | None, statut: str | None) -> list[dict]:
    query = "SELECT * FROM invoices WHERE statut_révision != ?"
    params: list = [STATUT_A_REVISER]
    if year:
        query += " AND exercice_fiscal = ?"
        params.append(year)
    if statut:
        query += " AND (statut_fiscal_profil = ? OR statut_fiscal_profil IS NULL)"
        params.append(statut)
    query += " ORDER BY date_document ASC"
    return [dict(r) for r in conn.execute(query, params).fetchall()]

# ── CSV ───────────────────────────────────────────────────────────────────────

CSV_COLS = [
    "date_document", "type_document", "numéro_facture",
    "émetteur_nom", "émetteur_siren", "émetteur_tva_intracom",
    "destinataire_nom",
    "montant_ht", "taux_tva", "montant_tva", "montant_ttc", "devise", "montant_eur",
    "catégorie", "déductible", "taux_déductibilité",
    "mode_paiement", "statut_paiement",
    "exercice_fiscal", "trimestre", "statut_fiscal_profil",
    "confiance", "fichier_source",
]

def write_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  CSV → {path}")

# ── XLSX ──────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT = Font(bold=True, color="F1F5F9", size=10)
_ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
_RED_FILL = PatternFill("solid", fgColor="FEE2E2")
_GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
_THIN = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN

def _autofit_columns(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 4, 40)

def _write_journal(wb, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Journal"
    headers = ["Date", "Type", "N° Facture", "Émetteur", "SIREN", "Montant HT",
               "TVA %", "TVA €", "Montant TTC", "Devise", "Catégorie",
               "Déductible", "Taux déd.", "Mode paiement", "Statut", "Fichier source"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, r in enumerate(rows, start=2):
        row_data = [
            r.get("date_document"), r.get("type_document"), r.get("numéro_facture"),
            r.get("émetteur_nom"), r.get("émetteur_siren"),
            r.get("montant_ht"), r.get("taux_tva"), r.get("montant_tva"), r.get("montant_ttc"),
            r.get("devise", "EUR"), r.get("catégorie"),
            "Oui" if r.get("déductible") else "Non",
            f"{int((r.get('taux_déductibilité') or 0) * 100)}%",
            r.get("mode_paiement"), r.get("statut_paiement"),
            r.get("fichier_source"),
        ]
        ws.append(row_data)
        fill = _ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=col)
            cell.border = _THIN
            if fill:
                cell.fill = fill
            if col == 12:
                cell.fill = _GREEN_FILL if r.get("déductible") else _RED_FILL

    _autofit_columns(ws)

def _write_recap(wb, rows: list[dict], statut: str | None) -> None:
    ws = wb.create_sheet("Récapitulatif")
    total_ttc = sum(r.get("montant_ttc") or 0 for r in rows)
    total_ht = sum(r.get("montant_ht") or 0 for r in rows)
    total_tva = sum(r.get("montant_tva") or 0 for r in rows)
    charges_ded = sum(
        (r.get("montant_ht") or r.get("montant_ttc") or 0) * (r.get("taux_déductibilité") or 0)
        for r in rows if r.get("déductible")
    )
    factures_emises = [r for r in rows if r.get("type_document") == "facture_émise"]
    ca = sum(r.get("montant_ttc") or 0 for r in factures_emises)

    data = [
        ["RÉCAPITULATIF FISCAL", ""],
        ["Statut fiscal", statut or "non défini"],
        ["", ""],
        ["DÉPENSES", ""],
        ["Total TTC dépenses", total_ttc],
        ["Total HT dépenses", total_ht],
        ["Total TVA", total_tva],
        ["Charges déductibles", charges_ded],
        ["", ""],
        ["RECETTES", ""],
        ["Chiffre d'affaires (factures émises)", ca],
        ["", ""],
        ["SOLDE", ""],
        ["Résultat net (CA - charges déd.)", ca - charges_ded],
    ]

    for row_data in data:
        ws.append(row_data)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    for row_idx in (1, 4, 10, 13):
        cell = ws.cell(row=row_idx, column=1)
        cell.font = Font(bold=True, color="1E293B")
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for row_idx in range(5, len(data) + 1):
        cell = ws.cell(row=row_idx, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00 €'

def _write_declaration(wb, rows: list[dict], year: int | None, statut: str | None) -> None:
    ws = wb.create_sheet("Déclaration")
    ws.append(["DONNÉES POUR DÉCLARATION FISCALE", ""])
    ws.append(["Exercice", year or "Tous"])
    ws.append(["Statut", statut or "non défini"])
    ws.append(["", ""])

    by_cat: dict[str, float] = {}
    for r in rows:
        if r.get("déductible"):
            cat = r.get("catégorie") or "autres"
            montant = (r.get("montant_ht") or r.get("montant_ttc") or 0) * (r.get("taux_déductibilité") or 1)
            by_cat[cat] = by_cat.get(cat, 0) + montant

    ws.append(["CHARGES PAR CATÉGORIE (déductibles)", "Montant €"])
    _style_header(ws, ws.max_row, 2)

    for cat, total in sorted(by_cat.items()):
        ws.append([cat, round(total, 2)])
        ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00 €'

    ws.append(["", ""])
    ws.append(["TOTAL CHARGES DÉDUCTIBLES", round(sum(by_cat.values()), 2)])
    ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00 €'
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

# ── Deadline calculées (URSSAF / DGFiP) ──────────────────────────────────────
# Vérifié 2026-05-05 — sources : URSSAF.fr, impots.gouv.fr, DGFiP

def _deadline_trimestrielle(quarter: int, year: int) -> date:
    # AE trimestriel : fin du mois suivant la fin du trimestre
    ends = {1: (year, 4, 30), 2: (year, 7, 31), 3: (year, 10, 31), 4: (year + 1, 1, 31)}
    year_end, month_end, day_end = ends[quarter]
    return date(year_end, month_end, day_end)

def _deadline_mensuelle_ae(month: int, year: int) -> date:
    # AE mensuel : fin du mois suivant
    next_month = month % 12 + 1
    next_year = year + (1 if month == 12 else 0)
    return date(next_year, next_month, calendar.monthrange(next_year, next_month)[1])

def _deadline_tva_mensuelle(month: int, year: int) -> date:
    # SASU/SARL TVA CA3 : 19 du mois suivant (télédéclaration)
    next_month = month % 12 + 1
    next_year = year + (1 if month == 12 else 0)
    return date(next_year, next_month, 19)

def _deadline_annuelle(year: int) -> date:
    # IR salarié : 31 mai de l'année suivante
    return date(year + 1, 5, 31)

def _period_label(cadence: str, period_key: str) -> str:
    if cadence == "trimestrielle":
        year, q = period_key.split("-Q")
        return f"T{q} {year}"
    if cadence == "mensuelle":
        year, month = period_key.split("-")
        return f"{MONTHS_FR_SHORT[int(month)]} {year}"
    return period_key  # annuelle

def _compute_deadlines(rows: list[dict], cadence: str, statut: str | None) -> list[tuple]:
    """Returns list of (label, ca, charges_ded, deadline_str)."""
    periods: dict[str, dict] = {}

    for r in rows:
        raw_date = r.get("date_document")
        if not raw_date:
            continue
        try:
            dt = datetime.fromisoformat(raw_date)
        except ValueError:
            continue

        if cadence == "trimestrielle":
            key = f"{dt.year}-Q{(dt.month - 1) // 3 + 1}"
        elif cadence == "mensuelle":
            key = f"{dt.year}-{dt.month:02d}"
        else:
            key = str(dt.year)

        period = periods.setdefault(key, {"ca": 0.0, "charges": 0.0})
        if r.get("type_document") == "facture_émise":
            period["ca"] += r.get("montant_ttc") or 0
        elif r.get("déductible"):
            period["charges"] += (r.get("montant_ht") or r.get("montant_ttc") or 0) * (r.get("taux_déductibilité") or 1)

    result = []
    for key in sorted(periods):
        period = periods[key]
        label = _period_label(cadence, key)

        if cadence == "trimestrielle":
            year, q = int(key.split("-Q")[0]), int(key.split("-Q")[1])
            dl = _deadline_trimestrielle(q, year)
        elif cadence == "mensuelle":
            year, month = int(key.split("-")[0]), int(key.split("-")[1])
            dl = (_deadline_tva_mensuelle(month, year) if statut in ("SASU", "SARL")
                  else _deadline_mensuelle_ae(month, year))
        else:
            year = int(key)
            dl = _deadline_annuelle(year)

        result.append((label, round(period["ca"], 2), round(period["charges"], 2), dl.strftime("%d/%m/%Y")))

    return result

def _write_stats(wb, rows: list[dict], year: int | None, statut: str | None, cadence: str) -> None:
    ws = wb.create_sheet("Statistiques")
    ws.cell(row=1, column=1).value = "STATISTIQUES"
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    # ── Bloc A : décompte de pièces par mois ──────────────────────────────────
    ws.append([])
    ws.append(["DÉCOMPTE DE PIÈCES PAR MOIS", "", "", "", "", "", ""])
    _style_header(ws, ws.max_row, 7)

    type_cols = {
        "facture_reçue": "Charges", "facture_émise": "Produits",
        "avoir_reçu": "Avoirs reçus", "avoir_émis": "Avoirs émis",
        "reçu": "Reçus", "note_de_frais": "Notes frais",
    }
    headers_a = ["Mois"] + list(type_cols.values()) + ["Total"]
    ws.append(headers_a)
    _style_header(ws, ws.max_row, len(headers_a))

    monthly_counts: dict[str, dict] = {}
    for r in rows:
        raw = r.get("date_document")
        if not raw:
            continue
        try:
            dt = datetime.fromisoformat(raw)
        except ValueError:
            continue
        key = f"{dt.year}-{dt.month:02d}"
        mc = monthly_counts.setdefault(key, {t: 0 for t in type_cols})
        mc[r.get("type_document", "facture_reçue")] = mc.get(r.get("type_document", "facture_reçue"), 0) + 1

    for key in sorted(monthly_counts):
        mc = monthly_counts[key]
        counts = [mc.get(t, 0) for t in type_cols]
        ws.append([key] + counts + [sum(counts)])

    # ── Bloc B : montants par mois ────────────────────────────────────────────
    ws.append([])
    ws.append(["MONTANTS PAR MOIS", "", "", ""])
    _style_header(ws, ws.max_row, 4)
    ws.append(["Mois", "Total charges HT", "CA HT", "Balance"])
    _style_header(ws, ws.max_row, 4)

    monthly_amounts: dict[str, dict] = {}
    for r in rows:
        raw = r.get("date_document")
        if not raw:
            continue
        try:
            dt = datetime.fromisoformat(raw)
        except ValueError:
            continue
        key = f"{dt.year}-{dt.month:02d}"
        ma = monthly_amounts.setdefault(key, {"charges": 0.0, "ca": 0.0})
        if r.get("type_document") == "facture_émise":
            ma["ca"] += r.get("montant_ht") or 0
        else:
            ma["charges"] += r.get("montant_ht") or r.get("montant_ttc") or 0

    for key in sorted(monthly_amounts):
        ma = monthly_amounts[key]
        balance = round(ma["ca"] - ma["charges"], 2)
        row_data = [key, round(ma["charges"], 2), round(ma["ca"], 2), balance]
        ws.append(row_data)
        for col in range(2, 5):
            ws.cell(row=ws.max_row, column=col).number_format = '#,##0.00 €'

    # ── Bloc C : périodes de déclaration ──────────────────────────────────────
    ws.append([])
    ws.append([f"PÉRIODES DE DÉCLARATION ({cadence.upper()})", "", "", ""])
    _style_header(ws, ws.max_row, 4)
    ws.append(["Période", "CA déclaré", "Charges déductibles", "Deadline indicative"])
    _style_header(ws, ws.max_row, 4)

    for label, ca, charges, deadline in _compute_deadlines(rows, cadence, statut):
        ws.append([label, ca, charges, deadline])
        ws.cell(row=ws.max_row, column=2).number_format = '#,##0.00 €'
        ws.cell(row=ws.max_row, column=3).number_format = '#,##0.00 €'

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

def write_xlsx(rows: list[dict], path: Path, year: int | None, statut: str | None, cadence: str = "trimestrielle") -> None:
    if not HAS_OPENPYXL:
        print("  [SKIP XLSX] openpyxl non installé — pip install openpyxl")
        return
    import io, zipfile as _zip
    wb = openpyxl.Workbook()
    _write_journal(wb, rows)
    _write_recap(wb, rows, statut)
    _write_declaration(wb, rows, year, statut)
    _write_stats(wb, rows, year, statut, cadence)
    _epoch = datetime(2000, 1, 1)
    wb.properties.created = _epoch
    wb.properties.modified = _epoch
    # Save to buffer then rewrite ZIP entries with fixed timestamps so
    # consecutive runs produce bit-identical files (openpyxl uses wall-clock time)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    _FIXED_DATE = (2000, 1, 1, 0, 0, 0)
    with _zip.ZipFile(path, "w", _zip.ZIP_DEFLATED) as zout:
        with _zip.ZipFile(buf) as zin:
            for entry in zin.infolist():
                entry.date_time = _FIXED_DATE
                zout.writestr(entry, zin.read(entry.filename))
    print(f"  XLSX → {path}")

# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Export carnet de compte CSV + XLSX")
    parser.add_argument("--year", type=int, help="Exercice fiscal (ex: 2025)")
    parser.add_argument("--statut", help="Profil fiscal (auto-entrepreneur, SASU, SARL, salarié)")
    parser.add_argument("--all", dest="all_years", action="store_true", help="Exporter toutes les années")
    parser.add_argument("--config", type=Path, default=Path("config.toml"))
    args = parser.parse_args()

    cfg = load_config(args.config)
    db_path = Path(cfg["paths"]["db"])
    output_dir = Path(cfg["paths"]["output"])
    year = None if args.all_years else (args.year or datetime.now().year)

    if not db_path.exists():
        print(f"Base introuvable : {db_path} — lance d'abord extract.py")
        return

    conn = open_db(db_path)

    profile = get_user_profile(conn)
    if profile is None:
        print("Profil non configuré. Lance le dashboard d'abord : python dashboard.py")
        sys.exit(1)

    statut = args.statut or profile["fiscal_profile"]
    cadence = profile["cadence"] or CADENCE_DEFAULTS.get(statut, "trimestrielle")

    rows = fetch_rows(conn, year, statut)
    conn.close()

    if not rows:
        print("Aucune donnée à exporter (vérifie l'année et le statut)")
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"-{year}" if year else "-all"
    stem = f"ledger{suffix}"

    for ext in (".csv", ".xlsx"):
        output_path = output_dir / f"{stem}{ext}"
        if output_path.exists():
            output_path.unlink()

    print(f"Export {len(rows)} entrées — exercice {year or 'tous'} — statut {statut} — cadence {cadence}")
    write_csv(rows, output_dir / f"{stem}.csv")
    write_xlsx(rows, output_dir / f"{stem}.xlsx", year, statut, cadence)

if __name__ == "__main__":
    main()
