"""Tests for export.py — CSV + XLSX generation."""

import csv
import sqlite3
import sys
from pathlib import Path

import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
import extract as ex
import export as exp
from tests.conftest import OVH_TEXT


def _insert(conn, overrides=None):
    base = ex.parse_invoice(OVH_TEXT, "test.pdf", "auto-entrepreneur")
    base["hash_fichier"] = "h_" + str(hash(str(overrides)))
    if overrides:
        base.update(overrides)
    ex.insert(conn, base)
    return base


# ── CSV ───────────────────────────────────────────────────────────────────────

class TestCSVExport:
    def test_csv_created(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn)
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=2026, statut="auto-entrepreneur")
        conn2.close()
        out = tmp_project / "output" / "ledger-2026.csv"
        exp.write_csv(rows, out)
        assert out.exists()

    def test_csv_has_expected_columns(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn)
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=None, statut=None)
        conn2.close()
        out = tmp_project / "output" / "ledger.csv"
        exp.write_csv(rows, out)
        reader = csv.DictReader(open(out))
        cols = reader.fieldnames
        for expected in ["date_document", "montant_ttc", "catégorie", "déductible", "fichier_source"]:
            assert expected in cols

    def test_csv_correct_values(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn)
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=None, statut=None)
        conn2.close()
        out = tmp_project / "output" / "ledger.csv"
        exp.write_csv(rows, out)
        row = list(csv.DictReader(open(out)))[0]
        assert float(row["montant_ttc"]) == 129.46
        assert row["catégorie"] == "hébergement"


# ── Year filter ───────────────────────────────────────────────────────────────

class TestYearFilter:
    def test_filter_by_year(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn, {"exercice_fiscal": 2025, "hash_fichier": "h1"})
        _insert(conn, {"exercice_fiscal": 2026, "hash_fichier": "h2"})
        conn.close()
        conn2 = exp.open_db(db_path)
        rows_2025 = exp.fetch_rows(conn2, year=2025, statut=None)
        rows_2026 = exp.fetch_rows(conn2, year=2026, statut=None)
        rows_all = exp.fetch_rows(conn2, year=None, statut=None)
        conn2.close()
        assert len(rows_2025) == 1
        assert len(rows_2026) == 1
        assert len(rows_all) == 2

    def test_excludes_à_réviser(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn, {"statut_révision": "validé", "hash_fichier": "h1"})
        _insert(conn, {"statut_révision": "à_réviser", "hash_fichier": "h2"})
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=None, statut=None)
        conn2.close()
        assert len(rows) == 1

    def test_empty_db_returns_empty(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        ex.open_db(db_path).close()
        conn = exp.open_db(db_path)
        rows = exp.fetch_rows(conn, year=None, statut=None)
        conn.close()
        assert rows == []


# ── XLSX ──────────────────────────────────────────────────────────────────────

class TestXLSXExport:
    def _make_xlsx(self, tmp_project, n=2):
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        for i in range(n):
            _insert(conn, {"hash_fichier": f"h{i}", "exercice_fiscal": 2026})
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=2026, statut="auto-entrepreneur")
        conn2.close()
        out = tmp_project / "output" / "ledger-2026.xlsx"
        exp.write_xlsx(rows, out, 2026, "auto-entrepreneur")
        return out

    def test_xlsx_created(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx(tmp_project)
        assert out.exists()

    def test_xlsx_four_sheets(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx(tmp_project)
        wb = openpyxl.load_workbook(out)
        assert set(wb.sheetnames) == {"Journal", "Récapitulatif", "Déclaration", "Statistiques"}

    def test_journal_has_data_rows(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx(tmp_project, n=3)
        wb = openpyxl.load_workbook(out)
        ws = wb["Journal"]
        # Row 1 = header, rows 2..N = pièces, dernière ligne = TOTAUX (livre-journal PCG).
        data_rows = [
            r for r in ws.iter_rows(min_row=2, values_only=True)
            if any(v for v in r) and r[5] != "TOTAUX"
        ]
        assert len(data_rows) == 3

    def test_recap_total_ttc(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx(tmp_project, n=2)
        wb = openpyxl.load_workbook(out)
        ws = wb["Récapitulatif"]
        values = {r[0]: r[1] for r in ws.iter_rows(values_only=True) if r[0] and r[1]}
        # 2 invoices × 129.46 TTC
        assert values.get("Total TTC dépenses") == pytest.approx(258.92, abs=0.01)

    def test_declaration_sheet_has_categories(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx(tmp_project, n=1)
        wb = openpyxl.load_workbook(out)
        ws = wb["Déclaration"]
        col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "hébergement" in col_a


# ── CLI main ──────────────────────────────────────────────────────────────────

def _patch_profile(monkeypatch, tmp_project):
    """Monkeypatch resolve_paths to point at tmp_project dirs."""
    import profiles as _profiles
    monkeypatch.setattr(_profiles, "resolve_paths", lambda slug: {
        "db":     tmp_project / "data" / "invoices.db",
        "output": tmp_project / "output",
    })


class TestExportCLI:
    def test_missing_db_prints_message(self, tmp_project, monkeypatch, capsys):
        _patch_profile(monkeypatch, tmp_project)
        (tmp_project / "data" / "invoices.db").unlink()
        import sys as _sys
        _sys.argv = ["export.py", "--profile", "test", "--year", "2025"]
        exp.main()
        out = capsys.readouterr().out
        assert "introuvable" in out

    def test_no_data_prints_message(self, tmp_project, monkeypatch, capsys):
        _patch_profile(monkeypatch, tmp_project)
        import sys as _sys
        _sys.argv = ["export.py", "--profile", "test", "--year", "2099"]
        exp.main()
        out = capsys.readouterr().out
        assert "Aucune" in out

    def test_generates_both_files(self, tmp_project, monkeypatch):
        _patch_profile(monkeypatch, tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn, {"exercice_fiscal": 2026})
        conn.close()
        import sys as _sys
        _sys.argv = ["export.py", "--profile", "test", "--year", "2026"]
        exp.main()
        assert (tmp_project / "output" / "ledger-2026.csv").exists()
        assert (tmp_project / "output" / "ledger-2026.xlsx").exists()

    def test_idempotent_double_run(self, tmp_project, monkeypatch):
        _patch_profile(monkeypatch, tmp_project)
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        _insert(conn, {"exercice_fiscal": 2026})
        conn.close()
        import sys as _sys
        _sys.argv = ["export.py", "--profile", "test", "--year", "2026"]
        exp.main()
        csv1 = (tmp_project / "output" / "ledger-2026.csv").read_bytes()
        exp.main()
        csv2 = (tmp_project / "output" / "ledger-2026.csv").read_bytes()
        assert csv1 == csv2


# ── Stats sheet ───────────────────────────────────────────────────────────────

class TestStatsSheet:
    def _make_xlsx_with_rows(self, tmp_project, rows_overrides):
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        for i, ov in enumerate(rows_overrides):
            _insert(conn, {"hash_fichier": f"hstat{i}", **ov})
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=None, statut=None)
        conn2.close()
        out = tmp_project / "output" / "ledger.xlsx"
        exp.write_xlsx(rows, out, None, "auto-entrepreneur", "trimestrielle")
        return out

    def test_stats_sheet_exists(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx_with_rows(tmp_project, [{"exercice_fiscal": 2025, "date_document": "2025-03-15"}])
        wb = openpyxl.load_workbook(out)
        assert "Statistiques" in wb.sheetnames

    def test_stats_has_deadline_bloc(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        out = self._make_xlsx_with_rows(tmp_project, [{"exercice_fiscal": 2025, "date_document": "2025-03-15"}])
        wb = openpyxl.load_workbook(out)
        ws = wb["Statistiques"]
        all_vals = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert any("DÉCLARATION" in str(v) for v in all_vals if v)

    def test_deadline_trimestrielle_q1(self):
        from export import _deadline_trimestrielle
        from datetime import date
        assert _deadline_trimestrielle(1, 2025) == date(2025, 4, 30)

    def test_deadline_trimestrielle_q4(self):
        from export import _deadline_trimestrielle
        from datetime import date
        assert _deadline_trimestrielle(4, 2025) == date(2026, 1, 31)

    def test_deadline_tva_mensuelle(self):
        from export import _deadline_tva_mensuelle
        from datetime import date
        assert _deadline_tva_mensuelle(1, 2025) == date(2025, 2, 19)
        assert _deadline_tva_mensuelle(12, 2025) == date(2026, 1, 19)

    def test_deadline_annuelle(self):
        from export import _deadline_annuelle
        from datetime import date
        assert _deadline_annuelle(2025) == date(2026, 5, 31)


# ── Assiette des charges selon le profil fiscal ─────────────────────────────
#
# Pour un auto-entrepreneur (franchise art. 293 B CGI) ou un salarié,
# la TVA payée aux fournisseurs n'est pas récupérable : l'assiette de
# la charge est TTC. Pour SASU/SARL, l'assiette historique reste HT.

def _charge_row(**overrides) -> dict:
    """Fixture d'une charge déductible à 100% — HT 100 / TVA 20 / TTC 120."""
    base = {
        "type_document": "facture_reçue",
        "montant_ht": 100.0,
        "montant_tva": 20.0,
        "montant_ttc": 120.0,
        "catégorie": "hébergement",
        "déductible": 1,
        "taux_déductibilité": 1.0,
        "date_document": "2025-03-15",
    }
    base.update(overrides)
    return base


def _read_recap_charges_deductibles(out_path: Path) -> float:
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Récapitulatif"]
    values = {r[0]: r[1] for r in ws.iter_rows(values_only=True) if r[0]}
    return values["Charges déductibles"]


def _read_declaration_total(out_path: Path) -> float:
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Déclaration"]
    for row in ws.iter_rows(values_only=True):
        if row[0] == "TOTAL CHARGES DÉDUCTIBLES":
            return row[1]
    raise AssertionError("Ligne TOTAL CHARGES DÉDUCTIBLES introuvable")


def _read_stats_block_b_headers(out_path: Path) -> list:
    # Bloc B = celui qui suit le titre « MONTANTS PAR MOIS ».
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Statistiques"]
    found_title = False
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val == "MONTANTS PAR MOIS":
            found_title = True
            continue
        if found_title and val == "Mois":
            return [ws.cell(row=r, column=c).value for c in range(1, 5)]
    raise AssertionError("Entêtes du bloc B introuvables")


def test_recapitulatif_charges_deductibles_auto_entrepreneur_en_ttc(tmp_path):
    # Given une charge déductible HT 100 / TTC 120
    rows = [_charge_row()]
    out = tmp_path / "ledger-ae.xlsx"
    # When on génère le XLSX pour un AE
    exp.write_xlsx(rows, out, 2025, "auto-entrepreneur", "trimestrielle")
    # Then la ligne « Charges déductibles » somme le TTC (120 € × 100%)
    assert _read_recap_charges_deductibles(out) == pytest.approx(120.0, abs=0.01)


def test_recapitulatif_charges_deductibles_sasu_en_ht_inchange(tmp_path):
    rows = [_charge_row()]
    out = tmp_path / "ledger-sasu.xlsx"
    exp.write_xlsx(rows, out, 2025, "SASU", "trimestrielle")
    # Then somme en HT (100 € × 100%) — non-régression
    assert _read_recap_charges_deductibles(out) == pytest.approx(100.0, abs=0.01)


def test_declaration_par_categorie_auto_entrepreneur_en_ttc(tmp_path):
    rows = [_charge_row(), _charge_row(catégorie="autres", montant_ht=50.0,
                                       montant_tva=10.0, montant_ttc=60.0)]
    out = tmp_path / "ledger-decl-ae.xlsx"
    exp.write_xlsx(rows, out, 2025, "auto-entrepreneur", "trimestrielle")
    # Then total catégoriel = 120 + 60 (TTC)
    assert _read_declaration_total(out) == pytest.approx(180.0, abs=0.01)


def test_declaration_par_categorie_sasu_en_ht_inchange(tmp_path):
    rows = [_charge_row(), _charge_row(catégorie="autres", montant_ht=50.0,
                                       montant_tva=10.0, montant_ttc=60.0)]
    out = tmp_path / "ledger-decl-sasu.xlsx"
    exp.write_xlsx(rows, out, 2025, "SASU", "trimestrielle")
    # Then total catégoriel = 100 + 50 (HT)
    assert _read_declaration_total(out) == pytest.approx(150.0, abs=0.01)


def test_statistiques_montants_par_mois_labels_ttc_pour_auto_entrepreneur(tmp_path):
    rows = [_charge_row()]
    out = tmp_path / "ledger-stats-ae.xlsx"
    exp.write_xlsx(rows, out, 2025, "auto-entrepreneur", "trimestrielle")
    headers = _read_stats_block_b_headers(out)
    # Le bandeau bascule en TTC pour rester cohérent avec l'assiette
    # comptable du profil
    assert headers == ["Mois", "Total charges TTC", "CA TTC", "Balance"]


def test_statistiques_montants_par_mois_labels_ht_pour_sasu(tmp_path):
    rows = [_charge_row()]
    out = tmp_path / "ledger-stats-sasu.xlsx"
    exp.write_xlsx(rows, out, 2025, "SASU", "trimestrielle")
    headers = _read_stats_block_b_headers(out)
    assert headers == ["Mois", "Total charges HT", "CA HT", "Balance"]


# ── Journal Débit/Crédit (livre-journal PCG) ─────────────────────────────────

class TestJournalDebitCredit:
    def _build_xlsx(self, tmp_project, rows_overrides):
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        for i, ov in enumerate(rows_overrides):
            _insert(conn, {"hash_fichier": f"hdc{i}", "exercice_fiscal": 2026, **ov})
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=2026, statut=None)
        conn2.close()
        out = tmp_project / "output" / "ledger-2026.xlsx"
        exp.write_xlsx(rows, out, 2026, "auto-entrepreneur")
        return openpyxl.load_workbook(out)["Journal"]

    @staticmethod
    def _headers(ws):
        return [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    @staticmethod
    def _data_rows(ws):
        # Toutes les lignes hors header et hors ligne TOTAUX.
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in r):
                continue
            if r[5] == "TOTAUX":
                continue
            rows.append(r)
        return rows

    def test_journal_a_les_colonnes_debit_credit_attendues(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [{"type_document": "facture_reçue"}])
        headers = self._headers(ws)
        for col in ("Débit HT", "Crédit HT", "Débit TVA", "Crédit TVA", "Débit TTC", "Crédit TTC"):
            assert col in headers

    def test_facture_recue_apparait_au_debit(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [{"type_document": "facture_reçue"}])
        headers = self._headers(ws)
        idx_debit_ht  = headers.index("Débit HT")
        idx_credit_ht = headers.index("Crédit HT")
        row = self._data_rows(ws)[0]
        assert row[idx_debit_ht] is not None
        assert row[idx_credit_ht] is None

    def test_facture_emise_apparait_au_credit(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [{
            "type_document": "facture_émise",
            "montant_ht": 1000.0, "montant_tva": 200.0, "montant_ttc": 1200.0,
        }])
        headers = self._headers(ws)
        idx_debit_ht  = headers.index("Débit HT")
        idx_credit_ht = headers.index("Crédit HT")
        row = self._data_rows(ws)[0]
        assert row[idx_debit_ht] is None
        assert row[idx_credit_ht] == 1000.0

    def test_avoir_recu_contre_passe_au_credit(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [{"type_document": "avoir_reçu"}])
        headers = self._headers(ws)
        row = self._data_rows(ws)[0]
        assert row[headers.index("Crédit HT")] is not None
        assert row[headers.index("Débit HT")] is None

    def test_avoir_emis_contre_passe_au_debit(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [{"type_document": "avoir_émis"}])
        headers = self._headers(ws)
        row = self._data_rows(ws)[0]
        assert row[headers.index("Débit HT")] is not None
        assert row[headers.index("Crédit HT")] is None

    def test_releve_bancaire_exclu_du_journal(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [
            {"type_document": "facture_reçue"},
            {"type_document": "relevé_bancaire"},
        ])
        # Seule la facture reçue doit apparaître.
        assert len(self._data_rows(ws)) == 1

    def test_totaux_journal_equilibres_pour_cycle_clos(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        # 1 facture émise + 1 facture reçue de même montant HT → équilibre.
        ws = self._build_xlsx(tmp_project, [
            {"type_document": "facture_reçue", "montant_ht": 500.0, "montant_tva": 100.0, "montant_ttc": 600.0},
            {"type_document": "facture_émise", "montant_ht": 500.0, "montant_tva": 100.0, "montant_ttc": 600.0},
        ])
        headers = self._headers(ws)
        # Trouver la ligne TOTAUX
        total_row = None
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r[5] == "TOTAUX":
                total_row = r
                break
        assert total_row is not None
        assert total_row[headers.index("Débit HT")] == pytest.approx(500.0)
        assert total_row[headers.index("Crédit HT")] == pytest.approx(500.0)

    def test_montant_none_donne_cellule_vide(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        ws = self._build_xlsx(tmp_project, [{
            "type_document": "facture_reçue",
            "montant_ht": 100.0, "montant_tva": None, "montant_ttc": None,
        }])
        headers = self._headers(ws)
        row = self._data_rows(ws)[0]
        assert row[headers.index("Débit HT")] == 100.0
        # TVA et TTC absents → cellule vide, pas un zéro parasite.
        assert row[headers.index("Débit TVA")] is None
        assert row[headers.index("Débit TTC")] is None


class TestCSVSensComptable:
    def _csv_rows(self, tmp_project, rows_overrides):
        db_path = tmp_project / "data" / "invoices.db"
        conn = ex.open_db(db_path)
        for i, ov in enumerate(rows_overrides):
            _insert(conn, {"hash_fichier": f"hsens{i}", **ov})
        conn.close()
        conn2 = exp.open_db(db_path)
        rows = exp.fetch_rows(conn2, year=None, statut=None)
        conn2.close()
        out = tmp_project / "output" / "ledger.csv"
        exp.write_csv(rows, out)
        return list(csv.DictReader(open(out)))

    def test_csv_inclut_colonne_sens_comptable(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        rows = self._csv_rows(tmp_project, [{"type_document": "facture_reçue"}])
        assert "sens_comptable" in rows[0]

    def test_facture_recue_sens_debit_dans_csv(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        rows = self._csv_rows(tmp_project, [{"type_document": "facture_reçue"}])
        assert rows[0]["sens_comptable"] == "débit"

    def test_facture_emise_sens_credit_dans_csv(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        rows = self._csv_rows(tmp_project, [{"type_document": "facture_émise"}])
        assert rows[0]["sens_comptable"] == "crédit"

    def test_releve_bancaire_sens_vide_dans_csv(self, tmp_project, monkeypatch):
        monkeypatch.chdir(tmp_project)
        rows = self._csv_rows(tmp_project, [{"type_document": "relevé_bancaire"}])
        # Ligne présente (CSV = registre plat) mais sens vide.
        assert rows[0]["sens_comptable"] == ""
